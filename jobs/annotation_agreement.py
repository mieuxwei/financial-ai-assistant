from __future__ import annotations

import argparse
import json
from pathlib import Path

from research.annotation.agreement import ReviewerAnnotation, build_agreement_report
from research.annotation.calibration_batch import ensure_private_output_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Merge completed reviewer workbooks and compute blinded agreement"
    )
    parser.add_argument("--reviewer-a", type=Path, required=True)
    parser.add_argument("--reviewer-b", type=Path, required=True)
    parser.add_argument("--reviewer-a-id", default="unknown")
    parser.add_argument("--reviewer-b-id", default="unknown")
    parser.add_argument("--reviewer-a-type", choices=("human", "ai", "unknown"), default="unknown")
    parser.add_argument("--reviewer-b-type", choices=("human", "ai", "unknown"), default="unknown")
    parser.add_argument("--minimum-kappa", type=float, default=0.60)
    parser.add_argument("--output", type=Path, required=True)
    return parser


def load_reviewer_workbook(path: Path) -> list[ReviewerAnnotation]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "workbook agreement requires optional dependency: pip install -e '.[annotation]'"
        ) from error

    workbook = load_workbook(path, read_only=True, data_only=False)
    if "Annotation" not in workbook.sheetnames:
        raise ValueError(f"missing Annotation sheet: {path}")
    sheet = workbook["Annotation"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    positions = {str(value): index for index, value in enumerate(headers) if value}
    required = {
        "candidate_id",
        "event_type",
        "impact_label",
        "confidence",
        "ambiguous_reason",
        "review_status",
        "exclusion_reason",
    }
    missing = sorted(required - positions.keys())
    if missing:
        raise ValueError(f"missing annotation columns in {path}: {missing}")

    output = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        candidate_id = values[positions["candidate_id"]]
        if not candidate_id:
            continue
        output.append(
            ReviewerAnnotation(
                candidate_id=str(candidate_id),
                event_type=str(values[positions["event_type"]] or ""),
                impact_label=str(values[positions["impact_label"]] or ""),
                confidence=int(values[positions["confidence"]] or 0),
                review_status=str(values[positions["review_status"]] or ""),
                ambiguous_reason=_optional_text(values[positions["ambiguous_reason"]]),
                exclusion_reason=_optional_text(values[positions["exclusion_reason"]]),
            )
        )
    return output


def _optional_text(value: object) -> str | None:
    text = str(value).strip() if value is not None else ""
    return text or None


def main() -> int:
    args = build_parser().parse_args()
    report = build_agreement_report(
        load_reviewer_workbook(args.reviewer_a),
        load_reviewer_workbook(args.reviewer_b),
        minimum_kappa=args.minimum_kappa,
    )
    report["reviewer_provenance"] = {
        "reviewer_a": {"id": args.reviewer_a_id, "type": args.reviewer_a_type},
        "reviewer_b": {"id": args.reviewer_b_id, "type": args.reviewer_b_type},
    }
    report["qualifies_as_two_human_review"] = (
        args.reviewer_a_type == "human" and args.reviewer_b_type == "human"
    )
    output = ensure_private_output_path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False))
    return 0 if report["passed_calibration_gate"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
